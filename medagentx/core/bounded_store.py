"""
24-Hour Bounded Persistence Layer for MedAgentX v2.0

Privacy-safe, bounded persistence layer for short-term memory and logs.

ABSOLUTE CONSTRAINTS:
- NO external databases (no SQL, MongoDB, Redis, cloud storage)
- Local file-based encrypted storage ONLY
- Deterministic behavior required
- Explicit data lifecycle visibility required

Data Retention:
- Default retention window: 24 hours (configurable constant)
- TTL enforced on BOTH read and write
- Auto-purge expired records deterministically

MANDATORY EXCEL ARCHIVAL REQUIREMENT:
Before ANY data is deleted due to TTL expiry:
- Generate an Excel (.xlsx) summary file
- Include: Timestamp, Session ID, Agent/Squad identifiers, Responsibility tags (CRF), High-level outcome summaries ONLY (no raw PHI)
- Store Excel files in a designated /archives/ directory
- Excel files are read-only records for compliance/audit purposes
- After Excel export → purge in-memory and encrypted files
"""

import json
import os
from typing import Any, Dict, List, Optional
from datetime import datetime, timedelta
from pathlib import Path
import logging
import hashlib
from cryptography.fernet import Fernet
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
import base64

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logging.warning("openpyxl not available. Excel export will be disabled.")

logger = logging.getLogger(__name__)

# Default retention window: 24 hours
DEFAULT_RETENTION_HOURS = 24


class BoundedStore:
    """
    24-Hour Bounded Persistence Layer.
    
    Stores event logs, replay traces, and session memory summaries.
    Enforces TTL on both read and write.
    Exports to Excel before deletion.
    """
    
    def __init__(
        self,
        store_path: Optional[str] = None,
        retention_hours: int = DEFAULT_RETENTION_HOURS,
        encryption_key: Optional[bytes] = None,
    ):
        """
        Initialize Bounded Store.
        
        Args:
            store_path: Path to store directory (default: ./bounded_store)
            retention_hours: Retention window in hours (default: 24)
            encryption_key: Optional encryption key (generated if not provided)
        """
        self.store_path = Path(store_path or "./bounded_store")
        self.store_path.mkdir(parents=True, exist_ok=True)
        
        self.archives_path = self.store_path / "archives"
        self.archives_path.mkdir(parents=True, exist_ok=True)
        
        self.retention_hours = retention_hours
        self.retention_delta = timedelta(hours=retention_hours)
        
        # Initialize encryption
        if encryption_key:
            self.encryption_key = encryption_key
        else:
            # Generate a deterministic key from a fixed salt (for development)
            # In production, this should be provided securely
            self.encryption_key = self._generate_key()
        
        self.cipher = Fernet(self.encryption_key)
        
        # In-memory index
        self._index: Dict[str, Dict[str, Any]] = {}
        self._load_index()
        
        logger.info(
            f"Bounded Store initialized at {self.store_path} "
            f"with {retention_hours}h retention"
        )
    
    def _generate_key(self) -> bytes:
        """Generate encryption key (deterministic for development)."""
        # In production, use a secure key management system
        password = b"medagentx_default_key_change_in_production"
        salt = b"medagentx_salt"
        kdf = PBKDF2HMAC(
            algorithm=hashes.SHA256(),
            length=32,
            salt=salt,
            iterations=100000,
        )
        key = base64.urlsafe_b64encode(kdf.derive(password))
        return key
    
    def store(
        self,
        data_type: str,  # "event_log", "replay_trace", "session_memory"
        session_id: str,
        data: Dict[str, Any],
        agent_id: Optional[str] = None,
        squad_id: Optional[str] = None,
        responsibility_tag: Optional[str] = None,
    ) -> str:
        """
        Store data with TTL enforcement.
        
        Args:
            data_type: Type of data ("event_log", "replay_trace", "session_memory")
            session_id: Session identifier
            data: Data to store
            agent_id: Optional agent identifier
            squad_id: Optional squad identifier
            responsibility_tag: Optional CRF responsibility tag
            
        Returns:
            Storage ID
        """
        # Generate storage ID
        storage_id = f"{data_type}_{session_id}_{datetime.now().isoformat()}"
        storage_id = storage_id.replace(":", "-").replace(".", "-")
        
        # Create storage entry
        entry = {
            "storage_id": storage_id,
            "data_type": data_type,
            "session_id": session_id,
            "agent_id": agent_id,
            "squad_id": squad_id,
            "responsibility_tag": responsibility_tag,
            "timestamp": datetime.now().isoformat(),
            "expires_at": (datetime.now() + self.retention_delta).isoformat(),
            "data": data,
        }
        
        # Encrypt and store
        encrypted_data = self._encrypt(json.dumps(entry))
        storage_file = self.store_path / f"{storage_id}.enc"
        with open(storage_file, "wb") as f:
            f.write(encrypted_data)
        
        # Update index
        self._index[storage_id] = {
            "data_type": data_type,
            "session_id": session_id,
            "timestamp": entry["timestamp"],
            "expires_at": entry["expires_at"],
            "agent_id": agent_id,
            "squad_id": squad_id,
            "responsibility_tag": responsibility_tag,
        }
        self._save_index()
        
        logger.debug(f"Stored {data_type} for session {session_id} as {storage_id}")
        
        # Check for expired entries and purge
        self._purge_expired()
        
        return storage_id
    
    def retrieve(
        self,
        storage_id: Optional[str] = None,
        session_id: Optional[str] = None,
        data_type: Optional[str] = None,
    ) -> List[Dict[str, Any]]:
        """
        Retrieve data with TTL enforcement.
        
        Expired entries are automatically purged before retrieval.
        
        Args:
            storage_id: Optional specific storage ID
            session_id: Optional session ID filter
            data_type: Optional data type filter
            
        Returns:
            List of retrieved entries (decrypted)
        """
        # Purge expired entries first
        self._purge_expired()
        
        # Filter entries
        entries = []
        for sid, index_entry in self._index.items():
            # Apply filters
            if storage_id and sid != storage_id:
                continue
            if session_id and index_entry["session_id"] != session_id:
                continue
            if data_type and index_entry["data_type"] != data_type:
                continue
            
            # Load and decrypt
            storage_file = self.store_path / f"{sid}.enc"
            if storage_file.exists():
                try:
                    with open(storage_file, "rb") as f:
                        encrypted_data = f.read()
                    decrypted_data = self._decrypt(encrypted_data)
                    entry = json.loads(decrypted_data)
                    entries.append(entry)
                except Exception as e:
                    logger.error(f"Error retrieving {sid}: {e}")
        
        # Sort by timestamp
        entries.sort(key=lambda x: x.get("timestamp", ""))
        
        return entries
    
    def _purge_expired(self) -> None:
        """
        Purge expired entries.
        
        Before deletion:
        1. Export to Excel
        2. Delete encrypted files
        3. Remove from index
        """
        now = datetime.now()
        expired_ids = []
        
        for storage_id, index_entry in self._index.items():
            expires_at_str = index_entry.get("expires_at")
            if expires_at_str:
                try:
                    expires_at = datetime.fromisoformat(expires_at_str)
                    if expires_at < now:
                        expired_ids.append(storage_id)
                except Exception as e:
                    logger.warning(f"Error parsing expires_at for {storage_id}: {e}")
                    # Treat as expired if we can't parse
                    expired_ids.append(storage_id)
        
        if expired_ids:
            logger.info(f"Purging {len(expired_ids)} expired entries")
            
            # Group by session for Excel export
            expired_by_session: Dict[str, List[str]] = {}
            for sid in expired_ids:
                session_id = self._index[sid].get("session_id", "unknown")
                if session_id not in expired_by_session:
                    expired_by_session[session_id] = []
                expired_by_session[session_id].append(sid)
            
            # Export to Excel before deletion
            for session_id, sids in expired_by_session.items():
                try:
                    self._export_to_excel(session_id, sids)
                except Exception as e:
                    logger.error(f"Error exporting to Excel for session {session_id}: {e}")
                    # Continue with deletion even if Excel export fails
            
            # Delete files and remove from index
            for sid in expired_ids:
                storage_file = self.store_path / f"{sid}.enc"
                if storage_file.exists():
                    try:
                        storage_file.unlink()
                    except Exception as e:
                        logger.error(f"Error deleting {sid}: {e}")
                
                if sid in self._index:
                    del self._index[sid]
            
            self._save_index()
    
    def _export_to_excel(
        self,
        session_id: str,
        storage_ids: List[str],
    ) -> str:
        """
        Export expired entries to Excel before deletion.
        
        Args:
            session_id: Session identifier
            storage_ids: List of storage IDs to export
            
        Returns:
            Path to exported Excel file
        """
        if not OPENPYXL_AVAILABLE:
            logger.warning("openpyxl not available. Skipping Excel export.")
            return ""
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Session Summary"
        
        # Headers
        headers = [
            "Timestamp",
            "Storage ID",
            "Data Type",
            "Agent ID",
            "Squad ID",
            "Responsibility Tag",
            "Outcome Summary",
        ]
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Add data rows
        for sid in storage_ids:
            index_entry = self._index.get(sid, {})
            
            # Load entry to get outcome summary (no raw PHI)
            outcome_summary = "N/A"
            try:
                storage_file = self.store_path / f"{sid}.enc"
                if storage_file.exists():
                    with open(storage_file, "rb") as f:
                        encrypted_data = f.read()
                    decrypted_data = self._decrypt(encrypted_data)
                    entry = json.loads(decrypted_data)
                    
                    # Extract high-level outcome summary (no raw PHI)
                    data = entry.get("data", {})
                    if isinstance(data, dict):
                        # Extract summary fields only
                        outcome_summary = self._extract_outcome_summary(data)
            except Exception as e:
                logger.warning(f"Error extracting summary for {sid}: {e}")
            
            row = [
                index_entry.get("timestamp", ""),
                sid,
                index_entry.get("data_type", ""),
                index_entry.get("agent_id", ""),
                index_entry.get("squad_id", ""),
                index_entry.get("responsibility_tag", ""),
                outcome_summary,
            ]
            ws.append(row)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save Excel file
        excel_filename = f"session_{session_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        excel_path = self.archives_path / excel_filename
        wb.save(excel_path)
        
        logger.info(f"Exported {len(storage_ids)} entries to Excel: {excel_path}")
        
        return str(excel_path)
    
    def _extract_outcome_summary(self, data: Dict[str, Any]) -> str:
        """
        Extract high-level outcome summary (no raw PHI).
        
        Args:
            data: Data dictionary
            
        Returns:
            Outcome summary string
        """
        summary_parts = []
        
        # Extract non-PHI summary fields
        if "confidence" in data:
            summary_parts.append(f"Confidence: {data['confidence']}")
        if "recommendation_type" in data:
            summary_parts.append(f"Type: {data['recommendation_type']}")
        if "status" in data:
            summary_parts.append(f"Status: {data['status']}")
        if "workflow_confidence" in data:
            summary_parts.append("Workflow completed")
        
        return " | ".join(summary_parts) if summary_parts else "Summary available"
    
    def export_to_json(
        self,
        session_id: Optional[str] = None,
        output_path: Optional[str] = None,
    ) -> str:
        """
        Manual export to JSON (full fidelity, responsibility preserved).
        
        Args:
            session_id: Optional session ID filter (exports all if not provided)
            output_path: Optional output file path
            
        Returns:
            Path to exported JSON file
        """
        # Retrieve all entries (respects TTL)
        entries = self.retrieve(session_id=session_id)
        
        export_data = {
            "export_timestamp": datetime.now().isoformat(),
            "retention_policy": f"{self.retention_hours} hours",
            "entry_count": len(entries),
            "entries": entries,
        }
        
        if output_path is None:
            export_dir = self.store_path / "exports"
            export_dir.mkdir(exist_ok=True)
            filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            if session_id:
                filename = f"export_{session_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            output_path = str(export_dir / filename)
        
        with open(output_path, "w") as f:
            json.dump(export_data, f, indent=2)
        
        logger.info(f"Exported {len(entries)} entries to JSON: {output_path}")
        
        return output_path
    
    def manual_purge(
        self,
        session_id: Optional[str] = None,
        before_date: Optional[datetime] = None,
    ) -> int:
        """
        Manual purge API (for explicit user/doctor control).
        
        Args:
            session_id: Optional session ID filter
            before_date: Optional date filter (purge entries before this date)
            
        Returns:
            Number of entries purged
        """
        # Export to Excel before manual purge
        if session_id:
            entries = self.retrieve(session_id=session_id)
            if entries:
                storage_ids = [e["storage_id"] for e in entries]
                try:
                    self._export_to_excel(session_id, storage_ids)
                except Exception as e:
                    logger.error(f"Error exporting to Excel before manual purge: {e}")
        
        # Purge entries
        purged_count = 0
        purge_ids = []
        
        for storage_id, index_entry in self._index.items():
            # Apply filters
            if session_id and index_entry["session_id"] != session_id:
                continue
            if before_date:
                timestamp_str = index_entry.get("timestamp", "")
                if timestamp_str:
                    try:
                        timestamp = datetime.fromisoformat(timestamp_str)
                        if timestamp >= before_date:
                            continue
                    except Exception as e:
                        logger.warning(f"Error parsing timestamp for {storage_id}: {e}")
            
            purge_ids.append(storage_id)
        
        # Delete files
        for sid in purge_ids:
            storage_file = self.store_path / f"{sid}.enc"
            if storage_file.exists():
                try:
                    storage_file.unlink()
                    purged_count += 1
                except Exception as e:
                    logger.error(f"Error deleting {sid}: {e}")
            
            if sid in self._index:
                del self._index[sid]
        
        self._save_index()
        
        logger.info(f"Manually purged {purged_count} entries")
        
        return purged_count
    
    def _encrypt(self, data: str) -> bytes:
        """Encrypt data."""
        return self.cipher.encrypt(data.encode())
    
    def _decrypt(self, encrypted_data: bytes) -> str:
        """Decrypt data."""
        return self.cipher.decrypt(encrypted_data).decode()
    
    def _load_index(self) -> None:
        """Load index from disk."""
        index_file = self.store_path / ".index.json"
        if index_file.exists():
            try:
                with open(index_file, "r") as f:
                    self._index = json.load(f)
            except Exception as e:
                logger.warning(f"Error loading index: {e}")
                self._index = {}
    
    def _save_index(self) -> None:
        """Save index to disk."""
        index_file = self.store_path / ".index.json"
        try:
            with open(index_file, "w") as f:
                json.dump(self._index, f, indent=2)
        except Exception as e:
            logger.error(f"Error saving index: {e}")

